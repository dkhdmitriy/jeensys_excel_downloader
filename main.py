import os
import sys
import time
import glob
from typing import Tuple, Dict, List
from collections import defaultdict

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# Optionally load environment variables from a local `.env` file if
# `python-dotenv` is installed. This makes local development easier while
# keeping secrets out of source control.
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    # Ignore if python-dotenv isn't installed; environment variables can
    # still be provided by the shell or CI environment.
    pass

SIGNIN_URL = "https://mon.jeensys.com/signin"
DEVICES_URL = "https://mon.jeensys.com/devices"
DEFAULT_TIMEOUT = 30


def read_credentials() -> Tuple[str, str]:
    # Read credentials from environment variables to avoid committing secrets.
    # Preferred variables: JEENSYS_LOGIN and JEENSYS_PASSWORD
    login = os.getenv("JEENSYS_LOGIN")
    password = os.getenv("JEENSYS_PASSWORD")

    if not login or not password:
        raise RuntimeError(
            "Credentials not found. Set environment variables JEENSYS_LOGIN and JEENSYS_PASSWORD "
            "or create a local `.env` file and load it before running. See `.env.example`."
        )

    return login, password


def build_driver(headless: bool = False) -> webdriver.Chrome:
    path = os.path.dirname(os.path.abspath(__file__))
    prefs = {'download.default_directory' : path}
    options = Options()
    options.add_experimental_option('prefs', prefs)
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--start-maximized")
    options.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36"
    )
    options.add_argument("--headless=new")
    options.add_argument("--window-size=1920,1080")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(60)
    return driver


def wait_for(driver: webdriver.Chrome, locator, timeout: int = DEFAULT_TIMEOUT):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located(locator))


def wait_clickable(driver: webdriver.Chrome, locator, timeout: int = DEFAULT_TIMEOUT):
    return WebDriverWait(driver, timeout).until(EC.element_to_be_clickable(locator))


def perform_login(driver: webdriver.Chrome, login: str, password: str) -> None:
    driver.get(SIGNIN_URL)

    email_input = wait_for(driver, (By.NAME, "username"))
    password_input = wait_for(driver, (By.NAME, "password"))

    email_input.clear()
    email_input.send_keys(login)

    password_input.clear()
    password_input.send_keys(password)

    submit_button = wait_for(driver, (By.CSS_SELECTOR, "button.authform__btn"))
    WebDriverWait(driver, 30).until(lambda d: submit_button.is_enabled())
    submit_button.click()

    WebDriverWait(driver, DEFAULT_TIMEOUT).until(
        EC.any_of(
            EC.url_contains("/dashboard"),
            EC.presence_of_element_located((By.CLASS_NAME, "ant-alert-error")),
        )
    )

    if driver.find_elements(By.CLASS_NAME, "ant-alert-error"):
        raise RuntimeError("Не удалось авторизоваться: проверьте логин и пароль.")


def open_devices_page(driver: webdriver.Chrome) -> None:
    driver.get(DEVICES_URL)
    wait_for(driver, (By.CSS_SELECTOR, "button.table__select"))


def select_all_devices(driver: webdriver.Chrome) -> None:
    select_locator = (By.CSS_SELECTOR, "button.table__select")
    wait_clickable(driver, select_locator).click()

    WebDriverWait(driver, DEFAULT_TIMEOUT).until(
        lambda d: "table__select-active"
        in d.find_element(*select_locator).get_attribute("class")
    )


def ensure_settings_panel_open(driver: webdriver.Chrome) -> None:
    panel_locator = (By.CSS_SELECTOR, ".settings.panel-box-blur")
    if driver.find_elements(*panel_locator):
        return

    toggle_locator = (By.CSS_SELECTOR, ".slot__buttons button:nth-child(2)")
    wait_clickable(driver, toggle_locator).click()
    wait_for(driver, panel_locator)


def start_excel_download(driver: webdriver.Chrome) -> None:
    export_button_locator = (
        By.XPATH,
        "//button[contains(@class,'settings__button') and .//span[text()='Скачать Excel']]",
    )
    wait_clickable(driver, export_button_locator).click()

    download_button_locator = (By.CSS_SELECTOR, "button.download__download")
    wait_clickable(driver, download_button_locator).click()

    confirm_button_locator = (By.CSS_SELECTOR, "button.confirmer__accept")
    wait_clickable(driver, confirm_button_locator).click()

    WebDriverWait(driver, DEFAULT_TIMEOUT).until(
        lambda d: not d.find_elements(*confirm_button_locator)
    )

    time.sleep(15)


def export_devices_excel(driver: webdriver.Chrome) -> None:
    open_devices_page(driver)
    select_all_devices(driver)
    ensure_settings_panel_open(driver)
    start_excel_download(driver)


def find_latest_excel() -> str:
    """Находит последний загруженный Excel файл в директории скрипта."""
    path = os.path.dirname(os.path.abspath(__file__))
    # Исключаем временные файлы Excel (начинающиеся с ~$)
    excel_files = [f for f in glob.glob(os.path.join(path, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if not excel_files:
        raise FileNotFoundError("Excel файл не найден в директории скрипта.")
    return max(excel_files, key=os.path.getctime)


def parse_excel_devices(excel_path: str) -> Dict[str, List[Dict]]:
    """
    Парсит Excel файл и разделяет устройства по клиентам.
    Применяет фильтр на каждого клиента, затем считает количество и хэшрейт.
    Название клиента в столбце I (индекс 8), хэшрейт в столбце O (индекс 14).
    Возвращает словарь: {клиент: [{name, hashrate}, ...]}
    """
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Маппинг названий клиентов для отчёта (точное совпадение!)
    client_mapping = {
        "Наша компания L7": "L7",
        "Наша компания L9": "L9",
        "Наша компания Whatsminer": "WM",
        "Наша компания T21": "T21",
        "Наша компания S19": "S19",
        "Наша Компания 2": "S19_dop",  # С большой буквой К
        "Наша компания EMCD": "S19_emcd",
    }
    
    devices_by_client = defaultdict(list)
    
    # Столбец I (индекс 8) - название клиента
    # Столбец O (индекс 14) - хэшрейт
    
    # Получаем все значения из столбца I для фильтрации
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    # Для каждого клиента из маппинга применяем фильтр с точным совпадением
    for full_name, short_name in client_mapping.items():
        # Фильтруем строки по точному названию клиента (без подстрок)
        filtered_rows = [
            row for row in all_rows 
            if row and len(row) > 8 and str(row[8]).strip() == full_name.strip()
        ]
        
        print(f"Фильтр '{full_name}': {len(filtered_rows)} устройств")
        
        # Для отфильтрованных строк считаем количество и хэшрейт
        for row in filtered_rows:
            if not row or len(row) < 9:
                continue
            
            client_name = row[8]  # Столбец I
            hashrate_str = row[14] if len(row) > 14 else "0"  # Столбец O
            
            if not client_name:
                continue
            
            # Парсим хэшрейт и нормализуем его (делим на 1000)
            try:
                hashrate_value = float(str(hashrate_str).replace(",", ".").split()[0])
                hashrate_value = hashrate_value / 1000  # Нормализуем значение
            except (ValueError, IndexError, AttributeError, TypeError):
                hashrate_value = 0.0
            
            devices_by_client[short_name].append({
                "name": client_name,
                "hashrate": hashrate_value,
            })
    
    return dict(devices_by_client)


def calculate_stats(devices: List[Dict]) -> Dict:
    """Рассчитывает статистику по устройствам."""
    if not devices:
        return {"count": 0, "total_hashrate": 0, "avg_hashrate": 0}
    
    total_hashrate = sum(d["hashrate"] for d in devices)
    avg_hashrate = total_hashrate / len(devices)
    
    return {
        "count": len(devices),
        "total_hashrate": total_hashrate,
        "avg_hashrate": avg_hashrate,
    }


def format_number(num: float, decimal_places: int = 2) -> str:
    """Форматирует число с запятой вместо точки."""
    if decimal_places == 0:
        # Округление до целого числа
        formatted = f"{round(num)}"
    else:
        # Округление до указанного числа знаков
        formatted = f"{num:.{decimal_places}f}"
    return formatted.replace(".", ",")


def generate_report(devices_by_client: Dict[str, List[Dict]]) -> None:
    """Генерирует и выводит отчёт по устройствам."""
    
    # Раздел LTC
    print("\n" + "="*50)
    print("ОТЧЁТ ПО УСТРОЙСТВАМ")
    print("="*50 + "\n")
    
    ltc_clients = {"L7": devices_by_client.get("L7", []),
                   "L9": devices_by_client.get("L9", [])}
    
    ltc_total_hashrate = 0

    for client_name, devices in ltc_clients.items():
        if not devices:
            continue
        
        stats = calculate_stats(devices)
        avg_hr = format_number(stats["avg_hashrate"], 2)
        count = int(round(stats["count"]))
        total_hr = int(round(stats["total_hashrate"]))
        
        print(f"{client_name}-{avg_hr} ({count}шт-{total_hr}хэш)")
        ltc_total_hashrate += stats["total_hashrate"]
    
    ltc_total_str = int(round(ltc_total_hashrate))
    print(f"ИТОГ LTC {ltc_total_str}\n")
    
    btc_clients_order = [
        ("WM", "WM"),
        ("T21", "T21"),
        ("S19", "S19"),
        ("S19_dop", "S19 доп"),
        ("S19_emcd", "S19 emcd"),
    ]
    
    s19_all_devices = []
    btc_total_hashrate = 0
    
    for client_key, display_name in btc_clients_order:
        devices = devices_by_client.get(client_key, [])
        
        if not devices:
            continue
        
        stats = calculate_stats(devices)
        avg_hr = format_number(stats["avg_hashrate"], 2)
        count = int(round(stats["count"]))
        total_hr = int(round(stats["total_hashrate"]))
        
        if client_key.startswith("S19"):
            s19_all_devices.extend(devices)
        
        # Вывод с комментариями для WM
        if client_key == "WM":
            print(f"{display_name}-{avg_hr} ({count}шт-{total_hr}хэш)")
        else:
            print(f"{display_name}-{avg_hr} ({count}шт-{total_hr}хэш)")
        
        btc_total_hashrate += stats["total_hashrate"]
    
    # Среднее по всем S19
    if s19_all_devices:
        s19_stats = calculate_stats(s19_all_devices)
        avg_s19 = format_number(s19_stats["avg_hashrate"], 2)
        s19_count = int(round(s19_stats["count"]))
        s19_total = int(round(s19_stats["total_hashrate"]))
        print(f"Среднее S19 - {avg_s19} ({s19_count}шт-{s19_total}хэш)")
    
    btc_total_str = int(round(btc_total_hashrate))
    print(f"ИТОГ BTC {btc_total_str}\n")


def analyze_and_report() -> None:
    """Находит Excel файл, анализирует данные и выводит отчёт."""
    try:
        excel_path = find_latest_excel()
        print(f"Анализирую файл: {excel_path}")
        
        devices_by_client = parse_excel_devices(excel_path)
        
        if not devices_by_client:
            print("Устройства не найдены в файле.")
            return
        
        generate_report(devices_by_client)
    
    except FileNotFoundError as e:
        print(f"Ошибка: {e}")
    except Exception as e:
        print(f"Ошибка при анализе файла: {e}")


def main():
    headless = "--headless" in sys.argv
    driver = build_driver(headless=headless)

    try:
        login, password = read_credentials()
        perform_login(driver, login, password)
        print("Авторизация завершена. Ожидайте экспорт устройств в Excel.")
        export_devices_excel(driver)
        print("Запрос на скачивание Excel отправлен.")
    except Exception as exc:
        print(f"Ошибка: {exc}")
    finally:
        if "--keep-open" in sys.argv:
            print("Оставляю браузер открытым на 15 секунд.")
            time.sleep(15)
        driver.quit()
    
    # После закрытия браузера анализируем Excel файл
    analyze_and_report()


if __name__ == "__main__":
    main()

